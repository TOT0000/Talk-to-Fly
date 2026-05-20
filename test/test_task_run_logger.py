import os
import json

import pytest

from controller.task_run_logger import TaskRunLogger, _OPENPYXL_AVAILABLE


@pytest.mark.skipif(not _OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_end_run_recovers_if_workbook_deleted(tmp_path):
    excel_path = tmp_path / "logs" / "task_runs.xlsx"
    logger = TaskRunLogger(excel_path=str(excel_path))

    logger.start_run(
        task_id="task-1",
        task_text="go to C1",
        scenario_name="scene-1",
        initial_snapshot={},
    )

    os.remove(excel_path)

    # Should not raise even if the workbook disappeared mid-run.
    logger.end_run(run_status="completed")

    assert excel_path.exists()


@pytest.mark.skipif(not _OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_start_run_autopersists_previous_pending_run(tmp_path):
    excel_path = tmp_path / "logs" / "task_runs.xlsx"
    logger = TaskRunLogger(excel_path=str(excel_path))

    logger.start_run(
        task_id="task-1",
        task_text="run one",
        scenario_name="scene-a",
        initial_snapshot={"benchmark_progress": {"completed": []}, "checkpoint_order": ["A1"]},
    )
    logger.end_run(run_status="completed")
    pending_summary = logger.get_pending_run_summary()
    first_run_id = str(pending_summary.get("run_id"))
    assert first_run_id

    logger.start_run(
        task_id="task-2",
        task_text="run two",
        scenario_name="scene-b",
        initial_snapshot={"benchmark_progress": {"completed": []}, "checkpoint_order": ["A1"]},
    )

    runtime_lines = (tmp_path / "logs" / "task_runs_runtime_trace.jsonl").read_text(encoding="utf-8").strip().splitlines()
    assert any(first_run_id in line for line in runtime_lines)
    assert (tmp_path / "logs" / f"{first_run_id}_summary.json").exists()


def test_planning_trace_filters_legacy_fields(tmp_path):
    excel_path = tmp_path / "logs" / "task_runs.xlsx"
    logger = TaskRunLogger(excel_path=str(excel_path))
    logger.start_run(
        task_id="task-legacy",
        task_text="legacy fields",
        scenario_name="scene-clean",
        initial_snapshot={"benchmark_progress": {"completed": []}, "checkpoint_order": ["A1"]},
    )
    logger.append_planning_trace(
        {
            "planning_stage": "initial",
            "llm_call_purpose": "initial",
            "prompt": "p",
            "raw_response": "r",
            "parsed_plan": "gc('A1');",
            "scene_id": "SCENE_X",
            "selected_baseline_id": "baseline1",
            "path_clear": False,
            "blocking_entity": "worker_1",
            "candidate_targets": [{"id": "A1"}],
            "generated_control_plan": "legacy",
        }
    )
    logger.end_run(run_status="completed")
    assert logger.save_pending_run() is True

    planning_lines = (tmp_path / "logs" / "task_runs_planning_trace.jsonl").read_text(encoding="utf-8").strip().splitlines()
    payload = json.loads(planning_lines[-1])
    assert "planning_stage" in payload
    assert "prompt" in payload
    assert "path_clear" not in payload
    assert "blocking_entity" not in payload
    assert "candidate_targets" not in payload
    assert "generated_control_plan" not in payload


def test_runtime_trace_remaining_checkpoints_are_active_scoped(tmp_path):
    excel_path = tmp_path / "logs" / "task_runs.xlsx"
    logger = TaskRunLogger(excel_path=str(excel_path))
    logger.start_run(
        task_id="task-active",
        task_text="active remaining",
        scenario_name="scene-active",
        initial_snapshot={
            "benchmark_progress": {"completed": ["A1"]},
            "checkpoint_order": ["A1", "B1", "C1"],
            "active_objective_set": {"active_checkpoint_ids": ["A1", "A2"]},
        },
    )
    logger.end_run(run_status="completed")
    assert logger.save_pending_run() is True

    runtime_lines = (tmp_path / "logs" / "task_runs_runtime_trace.jsonl").read_text(encoding="utf-8").strip().splitlines()
    payload = json.loads(runtime_lines[-1])
    assert payload["remaining_checkpoints"] == ["A2"]
    assert payload["global_unfinished_checkpoints"] == ["B1", "C1"]


def test_planning_trace_defaults_stage_and_source(tmp_path):
    excel_path = tmp_path / "logs" / "task_runs.xlsx"
    logger = TaskRunLogger(excel_path=str(excel_path))
    logger.start_run(
        task_id="task-plan-default",
        task_text="plan defaults",
        scenario_name="scene-plan",
        initial_snapshot={"benchmark_progress": {"completed": []}, "checkpoint_order": ["A1"]},
    )
    logger.append_planning_trace(
        {
            "llm_call_purpose": "heartbeat",
            "raw_response": "{\"response\":\"continue\"}",
            "current_target_checkpoint": "A1",
        }
    )
    logger.end_run(run_status="completed")
    assert logger.save_pending_run() is True
    planning_lines = (tmp_path / "logs" / "task_runs_planning_trace.jsonl").read_text(encoding="utf-8").strip().splitlines()
    payload = json.loads(planning_lines[-1])
    assert payload["planning_stage"] == "heartbeat"
    assert payload["plan_source"] == "heartbeat_decision"


def test_summary_completed_fields_follow_true_completion_state(tmp_path):
    excel_path = tmp_path / "logs" / "task_runs.xlsx"
    logger = TaskRunLogger(excel_path=str(excel_path))
    logger.start_run(
        task_id="task-true-completion",
        task_text="true completion sync",
        scenario_name="scene-sync",
        initial_snapshot={
            "benchmark_progress": {"completed": ["A1", "A2", "A3"]},
            "active_objective_set": {"active_checkpoint_ids": ["A1", "A2", "A4"]},
        },
    )
    logger.update_execution_info(
        execution_success=True,
        mission_success=False,
        termination_reason="queue_exhausted_with_unfinished_checkpoints",
        true_completed_checkpoints=["A1", "A2"],
        true_remaining_checkpoints=["A4"],
    )
    logger.end_run(run_status="incomplete")
    pending = logger.get_pending_run_summary()

    assert pending["completed_checkpoints"] == ["A1", "A2"]
    assert pending["true_completed_checkpoints"] == ["A1", "A2"]
    assert pending["completion_ratio"] == 2.0 / 3.0

@pytest.mark.skipif(not _OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_save_pending_run_persists_trajectory_fields_without_keyerror(tmp_path):
    from openpyxl import load_workbook

    excel_path = tmp_path / "logs" / "task_runs.xlsx"
    logger = TaskRunLogger(excel_path=str(excel_path))
    logger.start_run(
        task_id="task-trajectory",
        task_text="trajectory fields",
        scenario_name="scene-trajectory",
        initial_snapshot={"benchmark_progress": {"completed": []}, "checkpoint_order": ["A1"]},
    )
    logger.end_run(
        run_status="completed",
        final_snapshot={
            "benchmark_progress": {"completed": []},
            "checkpoint_order": ["A1"],
            "final_mission_summary": {
                "trajectory_sample_count": 12,
                "trajectory_buffer_source": "sampler",
                "trajectory_sampler_interval_sec": 0.25,
                "trajectory_sampler_active_during_run": True,
            },
        },
    )

    pending = logger.get_pending_run_summary()
    run_id = pending["run_id"]
    assert logger.save_pending_run() is True

    summary = json.loads((tmp_path / "logs" / "runs" / run_id / f"{run_id}_summary.json").read_text(encoding="utf-8"))
    assert summary["trajectory_sample_count"] == 12
    assert summary["trajectory_buffer_source"] == "sampler"
    assert summary["trajectory_sampler_interval_sec"] == 0.25
    assert summary["trajectory_sampler_active_during_run"] is True

    wb = load_workbook(excel_path)
    ws = wb["runs"]
    headers = [cell.value for cell in ws[1]]
    values = [cell.value for cell in ws[2]]
    row = dict(zip(headers, values))
    assert row["trajectory_sample_count"] == 12
    assert row["trajectory_buffer_source"] == "sampler"
    assert row["trajectory_sampler_interval_sec"] == 0.25
    assert row["trajectory_sampler_active_during_run"] is True


def test_save_pending_run_keeps_pending_if_persist_fails(tmp_path, monkeypatch):
    excel_path = tmp_path / "logs" / "task_runs.xlsx"
    logger = TaskRunLogger(excel_path=str(excel_path))
    logger.start_run(
        task_id="task-fail",
        task_text="persist fail",
        scenario_name="scene-fail",
        initial_snapshot={"benchmark_progress": {"completed": []}, "checkpoint_order": ["A1"]},
    )
    logger.end_run(run_status="completed")

    def _boom(_active):
        raise RuntimeError("persist boom")

    monkeypatch.setattr(logger, "_persist_run", _boom)

    with pytest.raises(RuntimeError, match="persist boom"):
        logger.save_pending_run()

    pending = logger.get_pending_run_summary()
    assert pending.get("run_id")
