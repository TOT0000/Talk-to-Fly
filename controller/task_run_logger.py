from __future__ import annotations

import csv
import json
import os
import threading
import time
from zipfile import BadZipFile
from dataclasses import asdict, dataclass, field, is_dataclass
from datetime import datetime, timezone
from typing import Dict, Optional, List
from uuid import uuid4

try:
    from openpyxl import Workbook, load_workbook
    _OPENPYXL_AVAILABLE = True
except Exception:
    Workbook = None
    load_workbook = None
    _OPENPYXL_AVAILABLE = False


RUNS_SHEET = "runs"
EVENTS_SHEET = "events"
DEBUG_SHEET = "debug"
DEFAULT_ARCHIVE_ROOT = "/home/jiafenli/typefly_logs_archive/archive/manual_runs"

RUN_COLUMNS = [
    "timestamp",
    "start_time",
    "end_timestamp",
    "run_id",
    "task_id",
    "task_text",
    "run_status",
    "scene_id",
    "baseline_scene_id",
    "selected_baseline_id",
    "selected_baseline_name",
    "framework_mode",
    "mission_success",
    "trigger_type",
    "trigger_params",
    "prompt_variant",
    "example_variant",
    "state_fields",
    "use_output_example",
    "archive_enabled",
    "saved_after_run",
    "task_success",
    "failure_reason",
    "completion_time_mission_sec",
    "total_replan_count",
    "total_llm_call_count",
    "actual_planning_call_count",
    "actual_evaluator_call_count",
    "planning_skipped_due_to_inflight_count",
    "evaluator_skipped_due_to_inflight_count",
    "planning_latency_mean_sec",
    "planning_latency_median_sec",
    "planning_latency_p90_sec",
    "planning_latency_p95_sec",
    "evaluator_latency_mean_sec",
    "evaluator_latency_median_sec",
    "evaluator_latency_p90_sec",
    "evaluator_latency_p95_sec",
    "all_llm_latency_mean_sec",
    "all_llm_latency_median_sec",
    "all_llm_latency_p90_sec",
    "all_llm_latency_p95_sec",
    "json_parse_success_rate",
    "collision_count",
    "near_miss_count",
    "min_uav_worker_distance_m",
    "completed_checkpoints",
    "completion_ratio",
    "planner_model_id",
    "evaluator_model_id",
    "heartbeat_seconds",
    "predicted_collision_threshold",
    "total_llm_wait_hover_sec",
    "completion_time_excluding_llm_wait_sec",
    "llm_wait_event_count",
    "lmstudio_base_url",
    "lmstudio_connected",
    "generated_plan",
    "final_plan_source",
]

EVENT_COLUMNS = [
    "run_id", "event_timestamp", "event_type", "details",
]

PLANNING_TRACE_ALLOWED_KEYS = {
    "planning_stage",
    "llm_call_purpose",
    "plan_source",
    "prompt_variant",
    "example_variant",
    "use_output_example",
    "prompt",
    "raw_response",
    "parsed_plan",
    "selected_baseline_id",
    "selected_baseline_name",
    "scene_id",
    "trigger_reason",
    "llm_called",
    "true_completed_checkpoints",
    "true_remaining_checkpoints",
    "current_target_checkpoint",
    "completion_state_source",
    "llm_call_id",
    "llm_call_role",
    "model_id",
    "request_start_ts",
    "response_end_ts",
    "latency_sec",
    "success",
    "response_type",
    "json_parse_success",
    "timeout",
    "skipped_due_to_inflight",
    "response_discarded_reason",
    "request_target_checkpoint",
}


@dataclass
class _RunRecord:
    run_id: str
    task_id: str
    task_text: str
    start_time: float
    start_iso: str
    run_status: str = "running"
    plan_generation_success: bool = False
    plan_execution_success: bool = False
    actual_plan_text: str = ""
    timeout_bool: bool = False
    failure_reason: str = ""
    task_completed_bool: bool = False

    initial_snapshot: Dict = field(default_factory=dict)
    final_snapshot: Dict = field(default_factory=dict)

    any_collision_during_run: bool = False
    _last_collision_state: bool = False

    baseline_info: Dict = field(default_factory=dict)
    planner_info: Dict = field(default_factory=dict)
    run_context: Dict = field(default_factory=dict)
    archive_enabled: bool = True
    results_only: bool = False
    runtime_trace: List[Dict] = field(default_factory=list)
    planning_trace: List[Dict] = field(default_factory=list)
    llm_call_count: int = 0
    near_miss_count: int = 0
    min_uav_worker_distance_m: Optional[float] = None
    mission_success: Optional[bool] = None
    termination_reason: str = ""
    queue_exhausted_with_unfinished: bool = False
    ended_due_to_replan_interrupt: bool = False
    true_completed_checkpoints: List[str] = field(default_factory=list)
    true_remaining_checkpoints: List[str] = field(default_factory=list)
    completion_state_source: str = "benchmark_progress/dwell_tracker"
    current_target_checkpoint: Optional[str] = None
    checkpoint_status_snapshot: Dict = field(default_factory=dict)
    completion_time_mission_sec: Optional[float] = None
    mission_start_ts: Optional[float] = None
    mission_end_ts: Optional[float] = None
    final_summary_source: str = "runtime_snapshot"
    final_runtime_snapshot_ts: Optional[str] = None
    final_completion_snapshot_ts: Optional[str] = None
    final_completion_snapshot_source: str = "benchmark_progress/dwell_tracker"
    final_execution_mode: str = ""
    final_run_status: str = ""
    final_mission_success: Optional[bool] = None
    final_termination_reason: str = ""
    final_true_completed_checkpoints: List[str] = field(default_factory=list)
    final_true_remaining_checkpoints: List[str] = field(default_factory=list)
    final_status_source: str = "final_mission_summary"
    interrupted_for_replan: bool = False
    entered_awaiting_replan_response: bool = False
    execution_resumed_from_new_plan: bool = False
    next_statement_executed_after_interrupt: bool = False


class TaskRunLogger:
    def __init__(self, excel_path: Optional[str] = None):
        archive_dir, resolved_excel_path = resolve_archive_root_and_excel_path(excel_path)
        self.excel_path = resolved_excel_path
        self.archive_dir = archive_dir
        self.runs_sheet_csv_path = resolve_runs_sheet_csv_path(self.excel_path)
        self.debug_jsonl_path = os.path.join(
            self.archive_dir,
            "task_runs_debug.jsonl",
        )
        self.runtime_trace_jsonl_path = os.path.join(
            self.archive_dir,
            "task_runs_runtime_trace.jsonl",
        )
        self.planning_trace_jsonl_path = os.path.join(
            self.archive_dir,
            "task_runs_planning_trace.jsonl",
        )
        self._lock = threading.Lock()
        self._active: Optional[_RunRecord] = None
        self._pending_completed: Optional[_RunRecord] = None
        self._enabled = _OPENPYXL_AVAILABLE
        self._warned_disabled = False
        self._ensure_archive_dirs()
        self._ensure_workbook()

    def _ensure_archive_dirs(self):
        os.makedirs(self.archive_dir, exist_ok=True)
        os.makedirs(os.path.join(self.archive_dir, "runs"), exist_ok=True)

    @staticmethod
    def _to_iso(ts: float) -> str:
        return datetime.fromtimestamp(float(ts), tz=timezone.utc).isoformat()

    @staticmethod
    def _json_text(value):
        if value is None:
            return ""
        if isinstance(value, str):
            return value

        def _default(o):
            if is_dataclass(o):
                return asdict(o)
            if hasattr(o, "to_dict") and callable(getattr(o, "to_dict")):
                try:
                    return o.to_dict()
                except Exception:
                    pass
            if hasattr(o, "__dict__"):
                try:
                    return {k: v for k, v in vars(o).items() if not str(k).startswith("_")}
                except Exception:
                    pass
            return str(o)

        return json.dumps(value, ensure_ascii=False, default=_default)

    def _ensure_workbook(self):
        if not self._enabled:
            self._warn_once_disabled()
            return
        os.makedirs(os.path.dirname(self.excel_path) or ".", exist_ok=True)
        if os.path.exists(self.excel_path):
            try:
                wb = load_workbook(self.excel_path)
            except BadZipFile:
                corrupted_path = f"{self.excel_path}.corrupted_{int(time.time())}"
                try:
                    os.replace(self.excel_path, corrupted_path)
                except Exception:
                    pass
                wb = Workbook()
                ws_runs = wb.active
                ws_runs.title = RUNS_SHEET
                ws_runs.append(RUN_COLUMNS)
                ws_events = wb.create_sheet(EVENTS_SHEET)
                ws_events.append(EVENT_COLUMNS)
                ws_debug = wb.create_sheet(DEBUG_SHEET)
                ws_debug.append(["run_id", "timestamp", "debug_json"])
                wb.save(self.excel_path)
                return
            self._ensure_sheet_schema(wb, RUNS_SHEET, RUN_COLUMNS)
            self._ensure_sheet_schema(wb, EVENTS_SHEET, EVENT_COLUMNS)
            self._ensure_sheet_schema(wb, DEBUG_SHEET, ["run_id", "timestamp", "debug_json"])
            wb.save(self.excel_path)
            return
        wb = Workbook()
        ws_runs = wb.active
        ws_runs.title = RUNS_SHEET
        ws_runs.append(RUN_COLUMNS)
        ws_events = wb.create_sheet(EVENTS_SHEET)
        ws_events.append(EVENT_COLUMNS)
        ws_debug = wb.create_sheet(DEBUG_SHEET)
        ws_debug.append(["run_id", "timestamp", "debug_json"])
        wb.save(self.excel_path)

    def _export_runs_sheet_csv(self, wb):
        csv_path = str(self.runs_sheet_csv_path or "").strip()
        if not csv_path:
            return
        if wb is None or RUNS_SHEET not in wb.sheetnames:
            return
        os.makedirs(os.path.dirname(csv_path) or ".", exist_ok=True)
        ws = wb[RUNS_SHEET]
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if value is None else value for value in row])

    def _load_workbook_resilient(self):
        if not self._enabled:
            return None
        try:
            return load_workbook(self.excel_path)
        except FileNotFoundError:
            self._ensure_workbook()
            return load_workbook(self.excel_path)
        except BadZipFile:
            self._ensure_workbook()
            return load_workbook(self.excel_path)

    def _ensure_sheet_schema(self, wb, sheet_name: str, expected_columns):
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(expected_columns)
            return
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
        if headers != list(expected_columns):
            del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
            ws.append(expected_columns)

    def start_run(self, task_id: str, task_text: str, scenario_name: str, initial_snapshot: Dict, archive_enabled: bool = True, run_context: Optional[Dict] = None):
        previous_pending = None
        with self._lock:
            if self._active is not None:
                return
            previous_pending = self._pending_completed
            self._pending_completed = None
            now = time.time()
            self._active = _RunRecord(
                run_id=f"run_{uuid4().hex[:12]}",
                task_id=task_id,
                task_text=task_text,
                start_time=now,
                start_iso=self._to_iso(now),
                initial_snapshot=initial_snapshot or {},
                archive_enabled=bool(archive_enabled),
                results_only=bool((run_context or {}).get("results_only", False)),
                run_context=dict(run_context or {}),
            )
            self._consume_snapshot(initial_snapshot, now=now)
        if previous_pending is not None:
            # Starting another WebUI run should not implicitly archive a previous
            # pending result. The user must explicitly choose Save this run or
            # Discard this run for completed manual runs.
            pass

    def update_plan_info(self, plan_text: str, generation_success: bool):
        with self._lock:
            if self._active is None:
                return
            self._active.actual_plan_text = plan_text or ""
            self._active.plan_generation_success = bool(generation_success)

    def update_execution_info(
        self,
        execution_success: bool,
        failure_reason: str = "",
        timeout_bool: bool = False,
        task_completed: bool = False,
        mission_success: Optional[bool] = None,
        termination_reason: str = "",
        queue_exhausted_with_unfinished: bool = False,
        ended_due_to_replan_interrupt: bool = False,
        true_completed_checkpoints: Optional[List[str]] = None,
        true_remaining_checkpoints: Optional[List[str]] = None,
        completion_state_source: str = "benchmark_progress/dwell_tracker",
        current_target_checkpoint: Optional[str] = None,
        checkpoint_status_snapshot: Optional[Dict] = None,
        completion_time_mission_sec: Optional[float] = None,
        mission_start_ts: Optional[float] = None,
        mission_end_ts: Optional[float] = None,
        final_summary_source: str = "runtime_snapshot",
        final_runtime_snapshot_ts: Optional[str] = None,
        final_completion_snapshot_ts: Optional[str] = None,
        final_completion_snapshot_source: str = "benchmark_progress/dwell_tracker",
        final_execution_mode: str = "",
        final_run_status: str = "",
        final_mission_success: Optional[bool] = None,
        final_termination_reason: str = "",
        final_true_completed_checkpoints: Optional[List[str]] = None,
        final_true_remaining_checkpoints: Optional[List[str]] = None,
        final_status_source: str = "final_mission_summary",
        interrupted_for_replan: bool = False,
        entered_awaiting_replan_response: bool = False,
        execution_resumed_from_new_plan: bool = False,
        next_statement_executed_after_interrupt: bool = False,
    ):
        with self._lock:
            if self._active is None:
                return
            self._active.plan_execution_success = bool(execution_success)
            self._active.timeout_bool = bool(timeout_bool)
            self._active.task_completed_bool = bool(task_completed)
            if mission_success is not None:
                self._active.mission_success = bool(mission_success)
            if termination_reason:
                self._active.termination_reason = str(termination_reason)
            self._active.queue_exhausted_with_unfinished = bool(queue_exhausted_with_unfinished)
            self._active.ended_due_to_replan_interrupt = bool(ended_due_to_replan_interrupt)
            if true_completed_checkpoints is not None:
                self._active.true_completed_checkpoints = [str(v).upper() for v in list(true_completed_checkpoints)]
            if true_remaining_checkpoints is not None:
                self._active.true_remaining_checkpoints = [str(v).upper() for v in list(true_remaining_checkpoints)]
            self._active.completion_state_source = str(completion_state_source or "benchmark_progress/dwell_tracker")
            self._active.current_target_checkpoint = (
                None if current_target_checkpoint is None else str(current_target_checkpoint).upper()
            )
            if checkpoint_status_snapshot is not None:
                self._active.checkpoint_status_snapshot = dict(checkpoint_status_snapshot)
            self._active.completion_time_mission_sec = (
                None if completion_time_mission_sec is None else float(completion_time_mission_sec)
            )
            if mission_start_ts is not None:
                self._active.mission_start_ts = float(mission_start_ts)
            if mission_end_ts is not None:
                self._active.mission_end_ts = float(mission_end_ts)
            self._active.final_summary_source = str(final_summary_source or "runtime_snapshot")
            self._active.final_runtime_snapshot_ts = final_runtime_snapshot_ts
            self._active.final_completion_snapshot_ts = final_completion_snapshot_ts
            self._active.final_completion_snapshot_source = str(final_completion_snapshot_source or "benchmark_progress/dwell_tracker")
            self._active.final_execution_mode = str(final_execution_mode or "")
            self._active.final_run_status = str(final_run_status or "")
            self._active.final_mission_success = None if final_mission_success is None else bool(final_mission_success)
            self._active.final_termination_reason = str(final_termination_reason or "")
            if final_true_completed_checkpoints is not None:
                self._active.final_true_completed_checkpoints = [str(v).upper() for v in list(final_true_completed_checkpoints)]
            if final_true_remaining_checkpoints is not None:
                self._active.final_true_remaining_checkpoints = [str(v).upper() for v in list(final_true_remaining_checkpoints)]
            self._active.final_status_source = str(final_status_source or "final_mission_summary")
            self._active.interrupted_for_replan = bool(interrupted_for_replan)
            self._active.entered_awaiting_replan_response = bool(entered_awaiting_replan_response)
            self._active.execution_resumed_from_new_plan = bool(execution_resumed_from_new_plan)
            self._active.next_statement_executed_after_interrupt = bool(next_statement_executed_after_interrupt)
            if failure_reason:
                self._active.failure_reason = str(failure_reason)

    def update_baseline_info(self, baseline_info: Dict):
        with self._lock:
            if self._active is None:
                return
            self._active.baseline_info = dict(baseline_info or {})

    def update_planner_info(self, planner_info: Dict):
        with self._lock:
            if self._active is None:
                return
            self._active.planner_info = dict(planner_info or {})

    def consume_runtime_snapshot(self, snapshot: Dict):
        with self._lock:
            if self._active is None:
                return
            self._consume_snapshot(snapshot, now=time.time())

    def append_planning_trace(self, trace: Dict):
        with self._lock:
            if self._active is None:
                return
            payload = dict(trace or {})
            if not payload.get("planning_stage"):
                purpose = str(payload.get("llm_call_purpose") or "").strip().lower()
                payload["planning_stage"] = (
                    "heartbeat" if "heartbeat" in purpose else ("replan" if "replan" in purpose else "initial")
                )
            if not payload.get("plan_source"):
                candidate_source = payload.get("plan_source") or payload.get("source") or payload.get("final_plan_source")
                if candidate_source:
                    payload["plan_source"] = str(candidate_source)
                else:
                    stage = str(payload.get("planning_stage") or "").strip().lower()
                    payload["plan_source"] = (
                        "heartbeat_decision" if stage == "heartbeat" else ("llm_replan" if stage == "replan" else "llm_initial")
                    )
            payload = {k: payload.get(k) for k in PLANNING_TRACE_ALLOWED_KEYS if k in payload}
            payload["run_id"] = self._active.run_id
            payload["timestamp"] = self._to_iso(time.time())
            if not bool(payload.get("skipped_due_to_inflight")):
                self._active.llm_call_count += 1
            self._active.planning_trace.append(payload)

    def _consume_snapshot(self, snapshot: Dict, now: float):
        if snapshot is None or self._active is None:
            return
        self._active.final_snapshot = snapshot
        self._active.runtime_trace.append(self._build_runtime_trace_row(snapshot, now))
        near_miss_count = int(snapshot.get("near_miss_count", 0) or 0)
        if near_miss_count > self._active.near_miss_count:
            self._active.near_miss_count = near_miss_count
        min_dist = snapshot.get("min_uav_worker_distance_m")
        if min_dist is not None:
            min_dist = float(min_dist)
            if self._active.min_uav_worker_distance_m is None or min_dist < self._active.min_uav_worker_distance_m:
                self._active.min_uav_worker_distance_m = min_dist
        collision_now = self._detect_collision(snapshot)
        self._active._last_collision_state = collision_now
        self._active.any_collision_during_run = self._active.any_collision_during_run or collision_now

    def _detect_collision(self, snapshot: Dict) -> bool:
        safety_context = snapshot.get("safety_context")
        if safety_context is None:
            return False
        return bool(getattr(safety_context, "envelopes_overlap", False))

    def _build_runtime_trace_row(self, snapshot: Dict, now: float) -> Dict:
        safety_context = snapshot.get("safety_context")
        benchmark_progress = dict(snapshot.get("benchmark_progress") or {})
        completed_checkpoints = [str(v).upper() for v in list(benchmark_progress.get("completed") or [])]
        checkpoint_order = [str(v).upper() for v in list(snapshot.get("checkpoint_order") or [])]
        global_unfinished_checkpoints = [cid for cid in checkpoint_order if cid not in set(completed_checkpoints)]
        active_checkpoint_ids = [
            str(v).upper()
            for v in list((snapshot.get("active_objective_set") or {}).get("active_checkpoint_ids") or [])
        ]
        remaining_checkpoints = [cid for cid in active_checkpoint_ids if cid not in set(completed_checkpoints)]
        return {
            "run_id": self._active.run_id if self._active else "",
            "timestamp": self._to_iso(now),
            "drone_gt": snapshot.get("drone_gt"),
            "drone_yaw_rad": snapshot.get("drone_yaw_rad"),
            "workers": snapshot.get("workers"),
            "benchmark_progress": benchmark_progress,
            "completed_checkpoints": completed_checkpoints,
            "remaining_checkpoints": remaining_checkpoints,
            "active_checkpoint_ids": active_checkpoint_ids,
            "global_unfinished_checkpoints": global_unfinished_checkpoints,
            "execution_mode": snapshot.get("execution_mode"),
            "framework_name": snapshot.get("framework_name"),
            "triggered_for_replan": bool(snapshot.get("replan_count", 0)),
            "replan_count": int(snapshot.get("replan_count", 0) or 0),
            "trigger_reason": snapshot.get("trigger_reason"),
            "predicted_collision_probability": None if safety_context is None else float(getattr(safety_context, "predicted_collision_probability", 0.0)),
            "per_worker_predicted_collision_probability": [] if safety_context is None else list(getattr(safety_context, "per_worker_collision_probabilities", []) or []),
            "dominant_risky_worker": None if safety_context is None else str(getattr(safety_context, "dominant_threat_id", "")),
            "near_miss_count": int(snapshot.get("near_miss_count", 0) or 0),
            "near_miss_events": list(snapshot.get("near_miss_events") or []),
            "collision_count": int(snapshot.get("collision_count", 0) or 0),
            "min_uav_worker_distance_m": snapshot.get("min_uav_worker_distance_m"),
            "scene_id": snapshot.get("baseline_scene_id"),
            "selected_baseline_id": snapshot.get("selected_baseline_id"),
            "completion_state_source": "benchmark_progress/dwell_tracker",
        }

    def _append_jsonl_line(self, path: str, payload: Dict):
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        with open(path, "a", encoding="utf-8") as f:
            f.write(self._json_text(payload) + "\n")

    def _write_json(self, path: str, payload: Dict):
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2, default=str)

    def _run_file_path(self, run_id: str, suffix: str) -> str:
        run_dir = os.path.join(self.archive_dir, "runs", str(run_id))
        os.makedirs(run_dir, exist_ok=True)
        return os.path.join(run_dir, f"{run_id}_{suffix}")

    def end_run(self, run_status: str, failure_reason: str = ""):
        with self._lock:
            if self._active is None:
                return
            active = self._active
            self._active = None
            active.run_status = run_status
            if failure_reason and not active.failure_reason:
                active.failure_reason = failure_reason
            self._pending_completed = active

    def has_pending_completed_run(self) -> bool:
        with self._lock:
            return self._pending_completed is not None

    @staticmethod
    def _percentile(sorted_values: List[float], percentile: float) -> Optional[float]:
        if not sorted_values:
            return None
        if len(sorted_values) == 1:
            return round(float(sorted_values[0]), 6)
        rank = (len(sorted_values) - 1) * float(percentile)
        low = int(rank)
        high = min(low + 1, len(sorted_values) - 1)
        weight = rank - low
        value = (sorted_values[low] * (1.0 - weight)) + (sorted_values[high] * weight)
        return round(float(value), 6)

    @classmethod
    def _latency_stats(cls, values: List[float], prefix: str) -> Dict:
        clean = sorted(float(v) for v in values if v is not None)
        if not clean:
            return {
                f"{prefix}_latency_mean_sec": None,
                f"{prefix}_latency_median_sec": None,
                f"{prefix}_latency_p90_sec": None,
                f"{prefix}_latency_p95_sec": None,
            }
        return {
            f"{prefix}_latency_mean_sec": round(float(sum(clean) / len(clean)), 6),
            f"{prefix}_latency_median_sec": cls._percentile(clean, 0.50),
            f"{prefix}_latency_p90_sec": cls._percentile(clean, 0.90),
            f"{prefix}_latency_p95_sec": cls._percentile(clean, 0.95),
        }

    @classmethod
    def _compute_llm_latency_summary(cls, active: _RunRecord) -> Dict:
        planning_roles = {"planning", "heartbeat", "replan", "initial_plan"}
        evaluator_roles = {"evaluator"}
        planning_latencies: List[float] = []
        evaluator_latencies: List[float] = []
        json_attempts = 0
        json_successes = 0
        actual_planning_count = 0
        actual_evaluator_count = 0
        planning_skips = 0
        evaluator_skips = 0
        for trace in list(active.planning_trace or []):
            role = str(trace.get("llm_call_role") or trace.get("planning_stage") or "").strip().lower()
            skipped = bool(trace.get("skipped_due_to_inflight"))
            if skipped:
                if role in evaluator_roles:
                    evaluator_skips += 1
                else:
                    planning_skips += 1
                continue
            is_actual = bool(trace.get("llm_call_id")) and trace.get("latency_sec") is not None
            if not is_actual:
                continue
            try:
                latency = float(trace.get("latency_sec"))
            except Exception:
                continue
            if role in evaluator_roles:
                evaluator_latencies.append(latency)
                actual_evaluator_count += 1
            elif role in planning_roles:
                planning_latencies.append(latency)
                actual_planning_count += 1
            if trace.get("json_parse_success") is not None:
                json_attempts += 1
                if bool(trace.get("json_parse_success")):
                    json_successes += 1
        all_latencies = list(planning_latencies) + list(evaluator_latencies)
        summary = {}
        summary.update(cls._latency_stats(planning_latencies, "planning"))
        summary.update(cls._latency_stats(evaluator_latencies, "evaluator"))
        summary.update(cls._latency_stats(all_latencies, "all_llm"))
        summary.update(
            {
                "actual_planning_call_count": int(actual_planning_count),
                "actual_evaluator_call_count": int(actual_evaluator_count),
                "planning_skipped_due_to_inflight_count": int(planning_skips),
                "evaluator_skipped_due_to_inflight_count": int(evaluator_skips),
                "actual_llm_request_count": int(actual_planning_count + actual_evaluator_count),
                "json_parse_success_rate": (None if json_attempts == 0 else round(float(json_successes / json_attempts), 6)),
            }
        )
        return summary

    def get_pending_run_summary(self) -> Dict:
        with self._lock:
            active = self._pending_completed
        if active is None:
            return {}
        final = active.final_snapshot or active.initial_snapshot or {}
        final_mission_summary = dict(final.get("final_mission_summary") or {})
        progress = dict(final.get("benchmark_progress") or {})
        objective_ids = [str(v).upper() for v in list((final.get("active_objective_set") or {}).get("active_checkpoint_ids") or [])]
        progress_completed = [str(v).upper() for v in list(progress.get("completed") or [])]
        true_completed = [str(v).upper() for v in list(active.true_completed_checkpoints or [])]
        if not true_completed:
            true_completed = [cid for cid in progress_completed if (not objective_ids) or (cid in objective_ids)]
        true_remaining = [str(v).upper() for v in list(active.true_remaining_checkpoints or [])]
        if not true_remaining:
            true_remaining = [cid for cid in objective_ids if cid not in set(true_completed)]
        active_scope_ids = list(objective_ids or sorted(set(true_completed + true_remaining)))
        completion_ratio = 0.0
        if active_scope_ids:
            completion_ratio = float(len([cid for cid in true_completed if cid in set(active_scope_ids)])) / float(len(active_scope_ids))
        llm_latency_summary = self._compute_llm_latency_summary(active)
        return {
            "run_id": active.run_id,
            "task_id": active.task_id,
            "selected_baseline_id": active.run_context.get("selected_baseline_id", ""),
            "selected_baseline_name": active.run_context.get("selected_baseline_name", ""),
            "framework_mode": active.run_context.get("framework_mode", ""),
            "scene_id": final.get("baseline_scene_id") or active.run_context.get("baseline_scene_id", ""),
            "run_status": active.run_status,
            "mission_success": active.mission_success,
            "termination_reason": active.termination_reason,
            "queue_exhausted_with_unfinished": bool(active.queue_exhausted_with_unfinished),
            "ended_due_to_replan_interrupt": bool(active.ended_due_to_replan_interrupt),
            "true_completed_checkpoints": list(true_completed),
            "true_remaining_checkpoints": list(true_remaining),
            "current_target_checkpoint": active.current_target_checkpoint,
            "checkpoint_status_snapshot": dict(active.checkpoint_status_snapshot),
            "completion_state_source": active.completion_state_source,
            "collision_count": int(final_mission_summary.get("collision_count", final.get("collision_count", 0)) or 0),
            "near_miss_count": int(final_mission_summary.get("near_miss_count", final.get("near_miss_count", 0)) or 0),
            "replan_count": int(final_mission_summary.get("replan_count", final.get("replan_count", 0)) or 0),
            "completion_time_mission_sec": active.completion_time_mission_sec,
            "mission_start_ts": active.mission_start_ts,
            "mission_end_ts": active.mission_end_ts,
            "final_summary_source": active.final_summary_source,
            "final_runtime_snapshot_ts": active.final_runtime_snapshot_ts,
            "final_completion_snapshot_ts": active.final_completion_snapshot_ts,
            "final_completion_snapshot_source": active.final_completion_snapshot_source,
            "final_execution_mode": active.final_execution_mode,
            "final_run_status": active.final_run_status,
            "final_mission_success": active.final_mission_success,
            "final_termination_reason": active.final_termination_reason,
            "final_true_completed_checkpoints": list(active.final_true_completed_checkpoints),
            "final_true_remaining_checkpoints": list(active.final_true_remaining_checkpoints),
            "final_status_source": active.final_status_source,
            "interrupted_for_replan": bool(active.interrupted_for_replan),
            "entered_awaiting_replan_response": bool(active.entered_awaiting_replan_response),
            "execution_resumed_from_new_plan": bool(active.execution_resumed_from_new_plan),
            "next_statement_executed_after_interrupt": bool(active.next_statement_executed_after_interrupt),
            "completed_checkpoints": list(true_completed),
            "completion_ratio": completion_ratio,
            "planner_model_id": active.run_context.get("planner_model_id", ""),
            "evaluator_model_id": active.run_context.get("evaluator_model_id", ""),
            "heartbeat_seconds": final_mission_summary.get("heartbeat_seconds", active.run_context.get("heartbeat_seconds")),
            "predicted_collision_threshold": final_mission_summary.get("predicted_collision_threshold", active.run_context.get("predicted_collision_threshold")),
            "total_llm_wait_hover_sec": float(final_mission_summary.get("total_llm_wait_hover_sec", 0.0) or 0.0),
            "completion_time_excluding_llm_wait_sec": float(final_mission_summary.get("completion_time_excluding_llm_wait_sec", active.completion_time_mission_sec or 0.0) or 0.0),
            "llm_wait_event_count": int(final_mission_summary.get("llm_wait_event_count", 0) or 0),
            "lmstudio_base_url": active.run_context.get("lmstudio_base_url", ""),
            "lmstudio_connected": bool(active.run_context.get("lmstudio_connected", False)),
            **llm_latency_summary,
            "runtime_trace_count": len(active.runtime_trace),
            "planning_trace_count": len(active.planning_trace),
        }

    def save_pending_run(self) -> bool:
        with self._lock:
            active = self._pending_completed
            self._pending_completed = None
        if active is None:
            return False
        self._persist_run(active)
        return True

    def discard_pending_run(self) -> bool:
        with self._lock:
            has_pending = self._pending_completed is not None
            self._pending_completed = None
        return has_pending

    def _persist_run(self, active: _RunRecord):
        end_ts = time.time()
        initial = active.initial_snapshot
        final = active.final_snapshot or initial or {}
        progress = dict(final.get("benchmark_progress") or {})
        objective_ids = [str(v).upper() for v in list((final.get("active_objective_set") or {}).get("active_checkpoint_ids") or [])]
        progress_completed = [str(v).upper() for v in list(progress.get("completed") or [])]
        true_completed = [str(v).upper() for v in list(active.true_completed_checkpoints or [])]
        if not true_completed:
            true_completed = [cid for cid in progress_completed if (not objective_ids) or (cid in objective_ids)]
        true_remaining = [str(v).upper() for v in list(active.true_remaining_checkpoints or [])]
        if not true_remaining:
            true_remaining = [cid for cid in objective_ids if cid not in set(true_completed)]
        active_scope_ids = list(objective_ids or sorted(set(true_completed + true_remaining)))
        completion_ratio = 0.0
        if active_scope_ids:
            completion_ratio = float(len([cid for cid in true_completed if cid in set(active_scope_ids)])) / float(len(active_scope_ids))

        persist_traces = not bool(active.results_only)
        if persist_traces:
            for row in active.runtime_trace:
                self._append_jsonl_line(self.runtime_trace_jsonl_path, row)
            for row in active.planning_trace:
                self._append_jsonl_line(self.planning_trace_jsonl_path, row)

        planner_info = active.planner_info or {}
        llm_latency_summary = self._compute_llm_latency_summary(active)
        final_mission_summary = dict(final.get("final_mission_summary") or {})
        completion_time_mission_sec = (
            None if not bool(active.mission_success)
            else (
                float(active.completion_time_mission_sec)
                if active.completion_time_mission_sec is not None
                else round(end_ts - active.start_time, 3)
            )
        )
        row = {
            "timestamp": active.start_iso,
            "start_time": active.start_iso,
            "end_timestamp": self._to_iso(end_ts),
            "run_id": active.run_id,
            "task_id": active.task_id,
            "task_text": active.task_text,
            "run_status": active.run_status,
            "scene_id": final.get("baseline_scene_id") or active.run_context.get("baseline_scene_id", ""),
            "baseline_scene_id": active.run_context.get("baseline_scene_id", ""),
            "selected_baseline_id": active.run_context.get("selected_baseline_id", ""),
            "selected_baseline_name": active.run_context.get("selected_baseline_name", ""),
            "framework_mode": active.run_context.get("framework_mode", ""),
            "mission_success": bool(active.mission_success),
            "trigger_type": active.run_context.get("trigger_type", ""),
            "trigger_params": self._json_text(active.run_context.get("trigger_params", {})),
            "prompt_variant": active.run_context.get("prompt_variant", ""),
            "example_variant": active.run_context.get("example_variant", ""),
            "state_fields": self._json_text(active.run_context.get("state_fields", [])),
            "use_output_example": active.run_context.get("use_output_example", ""),
            "archive_enabled": bool(active.run_context.get("archive_enabled", True)),
            "saved_after_run": True,
            "task_success": bool(active.task_completed_bool and active.plan_execution_success),
            "failure_reason": active.failure_reason,
            "completion_time_mission_sec": completion_time_mission_sec,
            "total_replan_count": int(final_mission_summary.get("replan_count", (final or {}).get("replan_count", 0)) or 0),
            "total_llm_call_count": int(active.llm_call_count),
            "actual_planning_call_count": llm_latency_summary["actual_planning_call_count"],
            "actual_evaluator_call_count": llm_latency_summary["actual_evaluator_call_count"],
            "planning_skipped_due_to_inflight_count": llm_latency_summary["planning_skipped_due_to_inflight_count"],
            "evaluator_skipped_due_to_inflight_count": llm_latency_summary["evaluator_skipped_due_to_inflight_count"],
            "planning_latency_mean_sec": llm_latency_summary["planning_latency_mean_sec"],
            "planning_latency_median_sec": llm_latency_summary["planning_latency_median_sec"],
            "planning_latency_p90_sec": llm_latency_summary["planning_latency_p90_sec"],
            "planning_latency_p95_sec": llm_latency_summary["planning_latency_p95_sec"],
            "evaluator_latency_mean_sec": llm_latency_summary["evaluator_latency_mean_sec"],
            "evaluator_latency_median_sec": llm_latency_summary["evaluator_latency_median_sec"],
            "evaluator_latency_p90_sec": llm_latency_summary["evaluator_latency_p90_sec"],
            "evaluator_latency_p95_sec": llm_latency_summary["evaluator_latency_p95_sec"],
            "all_llm_latency_mean_sec": llm_latency_summary["all_llm_latency_mean_sec"],
            "all_llm_latency_median_sec": llm_latency_summary["all_llm_latency_median_sec"],
            "all_llm_latency_p90_sec": llm_latency_summary["all_llm_latency_p90_sec"],
            "all_llm_latency_p95_sec": llm_latency_summary["all_llm_latency_p95_sec"],
            "json_parse_success_rate": llm_latency_summary["json_parse_success_rate"],
            "collision_count": int(final_mission_summary.get("collision_count", (final or {}).get("collision_count", 0)) or 0),
            "near_miss_count": int(final_mission_summary.get("near_miss_count", (final or {}).get("near_miss_count", 0)) or 0),
            "min_uav_worker_distance_m": active.min_uav_worker_distance_m,
            "completed_checkpoints": self._json_text(true_completed),
            "completion_ratio": completion_ratio,
            "planner_model_id": active.run_context.get("planner_model_id", planner_info.get("planner_model_id", "")),
            "evaluator_model_id": active.run_context.get("evaluator_model_id", planner_info.get("evaluator_model_id", "")),
            "heartbeat_seconds": final_mission_summary.get("heartbeat_seconds", active.run_context.get("heartbeat_seconds")),
            "predicted_collision_threshold": final_mission_summary.get("predicted_collision_threshold", active.run_context.get("predicted_collision_threshold")),
            "total_llm_wait_hover_sec": float(final_mission_summary.get("total_llm_wait_hover_sec", 0.0) or 0.0),
            "completion_time_excluding_llm_wait_sec": float(final_mission_summary.get("completion_time_excluding_llm_wait_sec", completion_time_mission_sec or 0.0) or 0.0),
            "llm_wait_event_count": int(final_mission_summary.get("llm_wait_event_count", 0) or 0),
            "lmstudio_base_url": active.run_context.get("lmstudio_base_url", ""),
            "lmstudio_connected": bool(active.run_context.get("lmstudio_connected", False)),
            "generated_plan": "" if bool(active.results_only) else active.actual_plan_text,
            "final_plan_source": planner_info.get("final_plan_source", ""),
        }
        run_summary = {
            "run_id": active.run_id,
            "start_time": active.start_iso,
            "end_time": self._to_iso(end_ts),
            "duration_sec": round(end_ts - active.start_time, 3),
            "run_status": active.run_status,
            "mission_success": active.mission_success,
            "termination_reason": active.termination_reason,
            "queue_exhausted_with_unfinished": bool(active.queue_exhausted_with_unfinished),
            "ended_due_to_replan_interrupt": bool(active.ended_due_to_replan_interrupt),
            "scene_id": final.get("baseline_scene_id") or active.run_context.get("baseline_scene_id", ""),
            "baseline_scene_id": active.run_context.get("baseline_scene_id", ""),
            "selected_baseline_id": active.run_context.get("selected_baseline_id", ""),
            "selected_baseline_name": active.run_context.get("selected_baseline_name", ""),
            "trigger_type": active.run_context.get("trigger_type", ""),
            "trigger_params": active.run_context.get("trigger_params", {}),
            "prompt_variant": active.run_context.get("prompt_variant", ""),
            "example_variant": active.run_context.get("example_variant", ""),
            "state_fields": active.run_context.get("state_fields", []),
            "use_output_example": bool(active.run_context.get("use_output_example", False)),
            "mission_start_ts": active.mission_start_ts,
            "mission_end_ts": active.mission_end_ts,
            "completion_time_mission_sec": completion_time_mission_sec,
            "replan_count": int(final_mission_summary.get("replan_count", (final or {}).get("replan_count", 0)) or 0),
            "llm_call_count": int(active.llm_call_count),
            **llm_latency_summary,
            "collision_count": int(final_mission_summary.get("collision_count", (final or {}).get("collision_count", 0)) or 0),
            "near_miss_count": int(final_mission_summary.get("near_miss_count", (final or {}).get("near_miss_count", 0)) or 0),
            "min_uav_worker_distance_m": active.min_uav_worker_distance_m,
            "completed_checkpoints": list(true_completed),
            "completion_ratio": completion_ratio,
            "true_completed_checkpoints": list(true_completed),
            "true_remaining_checkpoints": list(true_remaining),
            "current_target_checkpoint": active.current_target_checkpoint,
            "checkpoint_status_snapshot": dict(active.checkpoint_status_snapshot),
            "completion_state_source": active.completion_state_source,
            "final_summary_source": active.final_summary_source,
            "final_runtime_snapshot_ts": active.final_runtime_snapshot_ts,
            "final_completion_snapshot_ts": active.final_completion_snapshot_ts,
            "final_completion_snapshot_source": active.final_completion_snapshot_source,
            "final_execution_mode": active.final_execution_mode,
            "final_run_status": active.final_run_status,
            "final_mission_success": active.final_mission_success,
            "final_termination_reason": active.final_termination_reason,
            "final_true_completed_checkpoints": list(active.final_true_completed_checkpoints),
            "final_true_remaining_checkpoints": list(active.final_true_remaining_checkpoints),
            "final_status_source": active.final_status_source,
            "interrupted_for_replan": bool(active.interrupted_for_replan),
            "entered_awaiting_replan_response": bool(active.entered_awaiting_replan_response),
            "execution_resumed_from_new_plan": bool(active.execution_resumed_from_new_plan),
            "next_statement_executed_after_interrupt": bool(active.next_statement_executed_after_interrupt),
            "runtime_trace_count": len(active.runtime_trace),
            "planning_trace_count": len(active.planning_trace),
            "planner_model_id": row["planner_model_id"],
            "evaluator_model_id": row["evaluator_model_id"],
            "lmstudio_base_url": row["lmstudio_base_url"],
            "lmstudio_connected": bool(row["lmstudio_connected"]),
            "heartbeat_seconds": row["heartbeat_seconds"],
            "predicted_collision_threshold": row["predicted_collision_threshold"],
            "total_llm_wait_hover_sec": row["total_llm_wait_hover_sec"],
            "completion_time_excluding_llm_wait_sec": row["completion_time_excluding_llm_wait_sec"],
            "llm_wait_event_count": row["llm_wait_event_count"],
            "results_only": bool(active.results_only),
        }
        debug_payload = {
            "run_id": active.run_id,
            "timestamp": self._to_iso(end_ts),
            "run_summary": run_summary,
            "metrics": {
                "near_miss_count": int(active.near_miss_count),
                "min_uav_worker_distance_m": active.min_uav_worker_distance_m,
                "collision_count": int(final_mission_summary.get("collision_count", (final or {}).get("collision_count", 0)) or 0),
                "replan_count": int(final_mission_summary.get("replan_count", (final or {}).get("replan_count", 0)) or 0),
            },
            "runtime_trace_count": len(active.runtime_trace),
            "planning_trace_count": len(active.planning_trace),
            "results_only": bool(active.results_only),
        }

        summary_per_run_path = self._run_file_path(active.run_id, "summary.json")
        if persist_traces:
            runtime_per_run_path = self._run_file_path(active.run_id, "runtime_trace.jsonl")
            planning_per_run_path = self._run_file_path(active.run_id, "planning_trace.jsonl")
            for trace_row in active.runtime_trace:
                self._append_jsonl_line(runtime_per_run_path, trace_row)
            for trace_row in active.planning_trace:
                self._append_jsonl_line(planning_per_run_path, trace_row)
        self._write_json(summary_per_run_path, run_summary)
        if not bool(active.results_only):
            debug_per_run_path = self._run_file_path(active.run_id, "debug.json")
            self._write_json(debug_per_run_path, debug_payload)

        wb = self._load_workbook_resilient()
        if wb is not None:
            ws = wb[RUNS_SHEET]
            ws.append([row[col] for col in RUN_COLUMNS])
            if not bool(active.results_only):
                ws_debug = wb[DEBUG_SHEET]
                ws_debug.append([active.run_id, self._to_iso(end_ts), self._json_text(debug_payload)])
            wb.save(self.excel_path)
            self._export_runs_sheet_csv(wb)
        if not bool(active.results_only):
            self._append_jsonl_line(self.debug_jsonl_path, debug_payload)

    def _warn_once_disabled(self):
        if self._warned_disabled:
            return
        self._warned_disabled = True
        print(
            "[WARN] TaskRunLogger disabled because openpyxl is not installed. "
            "Install dependency with: pip install openpyxl"
        )


def resolve_runs_sheet_csv_path(excel_path: str) -> str:
    explicit_csv = str(os.getenv("TYPEFLY_TASK_LOG_CSV", "") or "").strip()
    if explicit_csv:
        return os.path.abspath(os.path.expanduser(explicit_csv))
    return os.path.join(os.path.dirname(os.path.abspath(excel_path)) or ".", "task_runs_runs_sheet.csv")


def resolve_archive_root_and_excel_path(explicit_excel_path: Optional[str] = None) -> tuple[str, str]:
    explicit_xlsx = str(explicit_excel_path or "").strip()
    env_xlsx = str(os.getenv("TYPEFLY_TASK_LOG_XLSX", "") or "").strip()
    env_root = str(os.getenv("TYPEFLY_ARCHIVE_ROOT", "") or "").strip()

    chosen_xlsx = explicit_xlsx or env_xlsx
    if chosen_xlsx:
        xlsx_path = os.path.abspath(os.path.expanduser(chosen_xlsx))
        archive_root = os.path.dirname(xlsx_path) or DEFAULT_ARCHIVE_ROOT
    else:
        archive_root = os.path.abspath(os.path.expanduser(env_root or DEFAULT_ARCHIVE_ROOT))
        xlsx_path = os.path.join(archive_root, "task_runs.xlsx")

    return archive_root, xlsx_path
