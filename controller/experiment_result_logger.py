from __future__ import annotations

import csv
import os
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Dict, Iterable

try:
    from openpyxl import Workbook, load_workbook
    _OPENPYXL_AVAILABLE = True
except Exception:
    Workbook = None
    load_workbook = None
    _OPENPYXL_AVAILABLE = False


RESULT_COLUMNS = [
    "timestamp",
    "experiment_tag",
    "pipeline_id",
    "scenario_id",
    "zone_id",
    "repeat_idx",
    "planner_model",
    "evaluator_model",
    "run_status",
    "mission_success",
    "termination_reason",
    "failure_reason",
    "completion_time_mission_sec",
    "replan_count",
    "collision_count",
    "near_miss_count",
    "min_uav_worker_distance_m",
    "completion_ratio",
    "completed_checkpoints",
    "remaining_checkpoints",
    "run_id",
    "task_id",
    "block_by",
    "block_model",
    "block_index",
]


@dataclass(frozen=True)
class ExperimentKey:
    planner_model: str
    evaluator_model: str
    repeat_idx: int


class ExperimentResultLogger:
    def __init__(self, csv_path: str, xlsx_path: str | None = None):
        self.csv_path = os.path.abspath(os.path.expanduser(csv_path))
        self.xlsx_path = (
            os.path.abspath(os.path.expanduser(xlsx_path))
            if xlsx_path
            else os.path.splitext(self.csv_path)[0] + ".xlsx"
        )
        os.makedirs(os.path.dirname(self.csv_path) or ".", exist_ok=True)
        self._ensure_csv_header()
        if _OPENPYXL_AVAILABLE:
            self._ensure_xlsx_header()

    def _ensure_csv_header(self):
        if os.path.exists(self.csv_path) and os.path.getsize(self.csv_path) > 0:
            return
        with open(self.csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=RESULT_COLUMNS)
            writer.writeheader()

    def _ensure_xlsx_header(self):
        if os.path.exists(self.xlsx_path):
            wb = load_workbook(self.xlsx_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
            if headers != RESULT_COLUMNS:
                wb.remove(ws)
                ws = wb.create_sheet("results")
                ws.append(RESULT_COLUMNS)
            wb.save(self.xlsx_path)
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "results"
        ws.append(RESULT_COLUMNS)
        wb.save(self.xlsx_path)

    def load_completed_keys(self) -> set[ExperimentKey]:
        keys: set[ExperimentKey] = set()
        if not os.path.exists(self.csv_path):
            return keys
        with open(self.csv_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    keys.add(
                        ExperimentKey(
                            planner_model=str(row.get("planner_model", "")).strip(),
                            evaluator_model=str(row.get("evaluator_model", "")).strip(),
                            repeat_idx=int(row.get("repeat_idx", "0") or 0),
                        )
                    )
                except Exception:
                    continue
        return keys

    def append_result(self, row: Dict[str, Any]):
        payload = {k: row.get(k, "") for k in RESULT_COLUMNS}
        payload["timestamp"] = payload.get("timestamp") or datetime.now(timezone.utc).isoformat()
        with open(self.csv_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=RESULT_COLUMNS)
            writer.writerow(payload)
        if _OPENPYXL_AVAILABLE:
            wb = load_workbook(self.xlsx_path)
            ws = wb.active
            ws.append([payload.get(c, "") for c in RESULT_COLUMNS])
            wb.save(self.xlsx_path)


def normalize_checkpoint_list(values: Iterable[Any]) -> str:
    return ";".join(str(v).upper() for v in values if str(v).strip())
